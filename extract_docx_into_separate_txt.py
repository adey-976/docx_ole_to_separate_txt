"""
Extract text from all embedded files in a .docx document into a separate .txt file.
Each embedded file's content is presented with contextual references (surrounding text,
location, filename, type) so an LLM can correlate the extracted content with the
original .docx document.

Usage:
    python3 extract_docx_into_separate_txt.py <path_to_docx> [output_txt_path]

If output_txt_path is not specified, output is saved as <input_name>_embedded_extractions.txt
"""

import zipfile
import os
import struct
import email
from xml.etree import ElementTree as ET
from io import BytesIO

import docx
import openpyxl
import olefile


NSMAP = {
    'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'o': 'urn:schemas-microsoft-com:office:office',
    'v': 'urn:schemas-microsoft-com:vml',
    'mc': 'http://schemas.openxmlformats.org/markup-compatibility/2006',
    'wp': 'http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
}

for prefix, uri in NSMAP.items():
    ET.register_namespace(prefix, uri)

PROGID_EXT_MAP = {
    'Word.Document.12': '.docx',
    'Word.Document.8': '.doc',
    'Excel.Sheet.12': '.xlsx',
    'Excel.Sheet.8': '.xls',
    'Excel.SheetMacroEnabled.12': '.xlsm',
    'Acrobat.Document.DC': '.pdf',
    'Acrobat.Document': '.pdf',
    'Package': '',
}

PROGID_TYPE_MAP = {
    'Word.Document.12': 'Word Document (.docx)',
    'Word.Document.8': 'Word Document (.doc)',
    'Excel.Sheet.12': 'Excel Spreadsheet (.xlsx)',
    'Excel.Sheet.8': 'Excel Spreadsheet (.xls)',
    'Excel.SheetMacroEnabled.12': 'Excel Macro-Enabled Workbook (.xlsm)',
    'Acrobat.Document.DC': 'PDF Document (.pdf)',
    'Acrobat.Document': 'PDF Document (.pdf)',
    'Package': 'Package',
}


def extract_xlsx_text(data: bytes) -> str:
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    lines = []
    for sheet in wb.sheetnames:
        lines.append(f"  [Sheet: {sheet}]")
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            row_text = '\t'.join(str(c) if c is not None else '' for c in row)
            if row_text.strip():
                lines.append(f"    {row_text}")
    wb.close()
    return '\n'.join(lines)


def extract_csv_text(data: bytes) -> str:
    try:
        return data.decode('utf-8')
    except UnicodeDecodeError:
        return data.decode('latin-1')


def extract_txt_text(data: bytes) -> str:
    try:
        return data.decode('utf-8')
    except UnicodeDecodeError:
        return data.decode('latin-1')


def extract_eml_text(data: bytes) -> str:
    try:
        text = data.decode('utf-8')
    except UnicodeDecodeError:
        text = data.decode('latin-1')
    try:
        msg = email.message_from_string(text)
        parts = []
        if msg['Subject']:
            parts.append(f"  Subject: {msg['Subject']}")
        if msg['From']:
            parts.append(f"  From: {msg['From']}")
        if msg['To']:
            parts.append(f"  To: {msg['To']}")
        if msg['Date']:
            parts.append(f"  Date: {msg['Date']}")
        body = msg.get_payload(decode=True)
        if body:
            parts.append(f"  Body: {body.decode('utf-8', errors='replace')}")
        elif isinstance(msg.get_payload(), str):
            parts.append(f"  Body: {msg.get_payload()}")
        return '\n'.join(parts)
    except Exception:
        return text


def extract_msg_text(data: bytes) -> str:
    try:
        import extract_msg
        msg = extract_msg.openMsg(BytesIO(data))
        parts = []
        if msg.subject:
            parts.append(f"  Subject: {msg.subject}")
        if msg.sender:
            parts.append(f"  From: {msg.sender}")
        if msg.to:
            parts.append(f"  To: {msg.to}")
        if msg.body:
            parts.append(f"  Body: {msg.body}")
        return '\n'.join(parts)
    except Exception:
        try:
            text = data.decode('utf-8')
        except UnicodeDecodeError:
            text = data.decode('latin-1')
        return text


def extract_pdf_text(data: bytes) -> str:
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(BytesIO(data))
        pages = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text()
            if text and text.strip():
                pages.append(f"  [Page {i+1}]")
                pages.append(f"    {text.strip()}")
        return '\n'.join(pages) if pages else "  [Empty PDF]"
    except ImportError:
        try:
            import pdfplumber
            with pdfplumber.open(BytesIO(data)) as pdf:
                pages = []
                for i, page in enumerate(pdf.pages):
                    text = page.extract_text()
                    if text and text.strip():
                        pages.append(f"  [Page {i+1}]")
                        pages.append(f"    {text.strip()}")
                return '\n'.join(pages) if pages else "  [Empty PDF]"
        except ImportError:
            return "  [PDF detected but PyPDF2/pdfplumber not installed]"


def extract_embedded_docx_text(data: bytes) -> str:
    doc = docx.Document(BytesIO(data))
    return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())


def extract_ole_embedded(data: bytes, prog_id: str = ''):
    try:
        ole = olefile.OleFileIO(BytesIO(data))
        if ole.exists('CONTENTS'):
            stream = ole.openstream('CONTENTS')
            file_data = stream.read()
            ole.close()
            ext = PROGID_EXT_MAP.get(prog_id, '')
            filename = f'embedded{ext}' if ext else 'embedded_contents'
            return filename, file_data
        if ole.exists('\x01Ole10Native'):
            stream = ole.openstream('\x01Ole10Native')
            content = stream.read()
            ole.close()
            return parse_ole10native(content)
        if ole.exists('Package'):
            stream = ole.openstream('Package')
            file_data = stream.read()
            ole.close()
            return 'unknown_package', file_data
        ole.close()
    except Exception:
        pass
    return None, None


def parse_ole10native(content: bytes):
    if len(content) < 8:
        return None, None
    idx = 0
    total_size = struct.unpack('<I', content[idx:idx+4])[0]
    idx += 4
    flags = struct.unpack('<H', content[idx:idx+2])[0]
    idx += 2
    end = content.index(b'\x00', idx)
    label = content[idx:end].decode('latin-1')
    idx = end + 1
    end = content.index(b'\x00', idx)
    src_path = content[idx:end].decode('latin-1')
    idx = end + 1
    filename = label if label else os.path.basename(src_path)
    if idx + 4 <= len(content):
        next_dword = struct.unpack('<I', content[idx:idx+4])[0]
        if next_dword == 0x00030000:
            idx += 4
            temp_len = struct.unpack('<I', content[idx:idx+4])[0]
            idx += 4
            idx += temp_len
            if idx + 4 <= len(content):
                data_size = struct.unpack('<I', content[idx:idx+4])[0]
                idx += 4
                file_data = content[idx:idx+data_size]
                return filename, file_data
        else:
            try:
                end = content.index(b'\x00', idx)
                idx = end + 1
                end = content.index(b'\x00', idx)
                idx = end + 1
                if idx + 4 <= len(content):
                    data_size = struct.unpack('<I', content[idx:idx+4])[0]
                    idx += 4
                    file_data = content[idx:idx+data_size]
                    return filename, file_data
            except ValueError:
                pass
    return None, None


def guess_extension_from_data(data: bytes) -> str:
    if data[:4] == b'PK\x03\x04':
        try:
            zf = zipfile.ZipFile(BytesIO(data))
            names = zf.namelist()
            if any('word/' in n for n in names):
                return '.docx'
            elif any('xl/' in n for n in names):
                return '.xlsx'
            elif any('ppt/' in n for n in names):
                return '.pptx'
            zf.close()
        except Exception:
            pass
        return '.zip'
    elif data[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        return '.ole'
    elif data[:5] == b'%PDF-':
        return '.pdf'
    elif data[:5] == b'MIME-' or data[:9] == b'Received:' or data[:5] == b'From:':
        return '.eml'
    return ''


def read_embedded_file(filename: str, data: bytes) -> str:
    ext = os.path.splitext(filename)[1].lower()
    if not ext:
        ext = guess_extension_from_data(data)
    if ext in ('.xlsx', '.xlsm'):
        return extract_xlsx_text(data)
    elif ext == '.csv':
        return extract_csv_text(data)
    elif ext == '.txt':
        return extract_txt_text(data)
    elif ext == '.eml':
        return extract_eml_text(data)
    elif ext == '.msg':
        return extract_msg_text(data)
    elif ext == '.docx':
        return extract_embedded_docx_text(data)
    elif ext == '.pdf':
        return extract_pdf_text(data)
    else:
        return f"  [Unsupported format: {ext or 'unknown'}]"


def get_paragraph_text(element):
    """Get plain text from a paragraph element."""
    return ''.join(
        node.text or ''
        for node in element.iter(f'{{{NSMAP["w"]}}}t')
    )


def get_surrounding_text(paragraphs, ole_para_index, max_chars=150):
    """Get text immediately before and after the OLE paragraph for context."""
    before_texts = []
    for i in range(ole_para_index - 1, -1, -1):
        t = get_paragraph_text(paragraphs[i]).strip()
        if t:
            before_texts.insert(0, t)
            if sum(len(x) for x in before_texts) >= max_chars:
                break
    after_texts = []
    for i in range(ole_para_index + 1, len(paragraphs)):
        t = get_paragraph_text(paragraphs[i]).strip()
        if t:
            after_texts.append(t)
            if sum(len(x) for x in after_texts) >= max_chars:
                break

    before = ' '.join(before_texts)
    after = ' '.join(after_texts)
    if len(before) > max_chars:
        before = '...' + before[-(max_chars - 3):]
    if len(after) > max_chars:
        after = after[:max_chars - 3] + '...'
    return before, after


def extract_embeddings(docx_path: str) -> str:
    """Extract all embedded files with contextual references."""
    zf = zipfile.ZipFile(docx_path)
    docx_filename = os.path.basename(docx_path)

    # Parse relationships
    rels_xml = zf.read('word/_rels/document.xml.rels')
    rels_tree = ET.fromstring(rels_xml)
    rels = {}
    rel_types = {}
    for rel in rels_tree:
        rid = rel.get('{http://schemas.openxmlformats.org/package/2006/relationships}Id')
        if rid is None:
            rid = rel.get('Id')
        target = rel.get('Target')
        rel_type = rel.get('Type', '')
        if rid and target:
            rels[rid] = target
            rel_types[rid] = rel_type

    # Parse document XML
    doc_xml = zf.read('word/document.xml')
    tree = ET.fromstring(doc_xml)
    body = tree.find(f'.//{{{NSMAP["w"]}}}body')

    w = NSMAP['w']
    o = NSMAP['o']
    r = NSMAP['r']

    embeddings = []
    embed_counter = 0

    def resolve_ole(obj):
        """Resolve an OLEObject element to (filename, file_data, prog_id)."""
        rid = obj.get(f'{{{r}}}id')
        if not rid or rid not in rels:
            return None, None, ''
        target = rels[rid]
        rel_type = rel_types.get(rid, '')
        prog_id = obj.get('ProgID', '')
        embed_path = f'word/{target}' if not target.startswith('/') else target.lstrip('/')
        try:
            embed_data = zf.read(embed_path)
        except KeyError:
            return None, None, prog_id

        filename = os.path.basename(target)
        file_data = None

        if 'package' in rel_type.lower():
            file_data = embed_data
            ext = os.path.splitext(filename)[1].lower()
            if not ext or ext == '.bin':
                prog_ext = PROGID_EXT_MAP.get(prog_id, '')
                if prog_ext:
                    filename = os.path.splitext(filename)[0] + prog_ext
        elif 'oleobject' in rel_type.lower().replace('/', ''):
            ole_filename, ole_data = extract_ole_embedded(embed_data, prog_id)
            if ole_filename and ole_data:
                filename = ole_filename
                file_data = ole_data
            else:
                file_data = embed_data
        else:
            ole_filename, ole_data = extract_ole_embedded(embed_data, prog_id)
            if ole_filename and ole_data:
                filename = ole_filename
                file_data = ole_data
            else:
                file_data = embed_data

        return filename, file_data, prog_id

    def process_paragraphs(paragraphs, location_prefix, override_index=None):
        """Process a list of paragraph elements, collecting embeddings with context.
        If override_index is set, only process the paragraph at that index but use
        the full paragraphs list for surrounding text context.
        """
        nonlocal embed_counter
        if override_index is not None:
            iter_items = [(override_index, paragraphs[override_index])]
        else:
            iter_items = list(enumerate(paragraphs))

        for i, para in iter_items:
            for obj in para.iter(f'{{{o}}}OLEObject'):
                filename, file_data, prog_id = resolve_ole(obj)
                if not filename or not file_data:
                    continue

                embed_counter += 1
                before, after = get_surrounding_text(paragraphs, i)
                file_type = PROGID_TYPE_MAP.get(prog_id, '')
                if not file_type:
                    ext = os.path.splitext(filename)[1].lower()
                    if not ext:
                        ext = guess_extension_from_data(file_data)
                    file_type = ext.lstrip('.').upper() + ' file' if ext else 'Unknown'

                content = read_embedded_file(filename, file_data)

                embeddings.append({
                    'number': embed_counter,
                    'filename': filename,
                    'file_type': file_type,
                    'location': location_prefix,
                    'text_before': before,
                    'text_after': after,
                    'content': content,
                })

    # Collect all body-level paragraphs (for context around standalone OLE paragraphs)
    body_elements = list(body)
    body_paragraphs = []
    body_para_index_map = {}
    for idx, el in enumerate(body_elements):
        el_tag = el.tag.split('}')[-1] if '}' in el.tag else el.tag
        if el_tag == 'p':
            body_para_index_map[idx] = len(body_paragraphs)
            body_paragraphs.append(el)

    # Walk through body elements
    body_element_index = 0
    for element in body_elements:
        tag = element.tag.split('}')[-1] if '}' in element.tag else element.tag
        body_element_index += 1

        if tag == 'p':
            ole_objects = list(element.iter(f'{{{o}}}OLEObject'))
            if ole_objects:
                para_idx = body_para_index_map.get(body_element_index - 1, 0)
                process_paragraphs(body_paragraphs, f"Body paragraph {body_element_index}",
                                   override_index=para_idx)

        elif tag == 'tbl':
            row_index = 0
            for row in element.iter(f'{{{w}}}tr'):
                row_index += 1
                cell_index = 0
                for cell in row.iter(f'{{{w}}}tc'):
                    cell_index += 1
                    ole_in_cell = cell.findall(f'.//{{{o}}}OLEObject')
                    if ole_in_cell:
                        paras = list(cell.iter(f'{{{w}}}p'))
                        location = f"Table {body_element_index}, Row {row_index}, Cell {cell_index}"
                        process_paragraphs(paras, location)

    zf.close()

    # Format output
    output = []
    output.append(f"{'='*80}")
    output.append(f"EMBEDDED FILE EXTRACTIONS FROM: {docx_filename}")
    output.append(f"{'='*80}")
    output.append("")
    output.append("This file contains the extracted text content of all embedded files found in the")
    output.append(f"document \"{docx_filename}\". Each extraction below includes the embedded file's")
    output.append("name, type, location in the document, and the surrounding text from the original")
    output.append("document so you can identify exactly where each embedded file appears.")
    output.append("")
    output.append(f"Total embedded files found: {len(embeddings)}")
    output.append("")

    for emb in embeddings:
        output.append(f"{'─'*80}")
        output.append(f"EMBEDDED FILE #{emb['number']}")
        output.append(f"{'─'*80}")
        output.append(f"  Filename : {emb['filename']}")
        output.append(f"  Type     : {emb['file_type']}")
        output.append(f"  Location : {emb['location']}")
        output.append(f"")
        if emb['text_before']:
            output.append(f"  Text immediately before this embedded file in the document:")
            output.append(f"    \"{emb['text_before']}\"")
            output.append("")
        if emb['text_after']:
            output.append(f"  Text immediately after this embedded file in the document:")
            output.append(f"    \"{emb['text_after']}\"")
            output.append("")
        output.append(f"  Extracted content:")
        output.append(f"  {'·'*40}")
        output.append(emb['content'])
        output.append(f"  {'·'*40}")
        output.append("")

    output.append(f"{'='*80}")
    output.append("END OF EXTRACTIONS")
    output.append(f"{'='*80}")

    return '\n'.join(output)


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print("Usage: python3 extract_docx_into_separate_txt.py <path_to_docx> [output_txt_path]")
        sys.exit(1)

    input_path = sys.argv[1]
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        base = os.path.splitext(input_path)[0]
        output_path = f"{base}_embedded_extractions.txt"

    result = extract_embeddings(input_path)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(result)
    print(f"Output saved to: {output_path}")
